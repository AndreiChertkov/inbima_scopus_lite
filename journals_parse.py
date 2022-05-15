import openpyxl
from time import perf_counter as tpc


from excel import Excel
from journals_description import description


CODES_MAX = 10000
JOURNALS_MAX = 10000000


def build(fpath='./data/scopus.xlsx'):
    journals = {}

    wb = openpyxl.load_workbook(fpath)

    sh = wb['codes']
    codes = _parse_codes(sh)

    sh = wb['data']
    ids = []
    for i in range(2, JOURNALS_MAX):
        id = sh.cell(i, 1).value
        if not id in ids:
            ids.append(id)

        title = sh.cell(i, 2).value

        if not id or not title:
            break

        error = ''

        issn = str(sh.cell(i, 20).value or '')
        eissn = str(sh.cell(i, 21).value or '')
        if not eissn and len(issn) > 8 and ' ' in issn:
            # Bug for some records (issn and eissn are in the same cell):
            eissn = issn.split(' ')[1]
            issn = issn.split(' ')[0]
        issn = _parse_issn(issn)
        eissn = _parse_issn(eissn)
        if not issn and not eissn:
            error += _err('ISSN and EISSN are empty [skip]', title)
            continue

        numb = sh.cell(i, 9).value
        subj = codes[numb]['full']
        q = f'Q{sh.cell(i, 17).value}'
        if not q in ['Q1', 'Q2', 'Q3', 'Q4']:
            error += _err('Invalid quartile [skip]', title)
            continue

        q1 = [subj] if q == 'Q1' else []
        q2 = [subj] if q == 'Q2' else []
        q3 = [subj] if q == 'Q3' else []
        q4 = [subj] if q == 'Q4' else []

        journal_ref = journals.get(id)
        if journal_ref:
            if title != journal_ref['title']:
                text = _err('Invalid id/title pair [skip]', title)
                journal_ref['error'] = '; '.join([journal_ref['error'], text])
            else:
                journal_ref['q1'].extend(q1)
                journal_ref['q2'].extend(q2)
                journal_ref['q3'].extend(q3)
                journal_ref['q4'].extend(q4)
                journal_ref['quartile'] = _join_q(
                    _get_q(q1, q2, q3, q4), journal_ref['quartile'])
            continue

        is_repeated_title = False
        for journal in journals.values():
            if journal['title'] == title:
                text = _err('Title repeated [skip]', title)
                journal['error'] = '; '.join([journal['error'], text])
                is_repeated_title = True
        if is_repeated_title:
            continue

        journals[id] = {
            'title': title,
            'issn': issn,
            'eissn': eissn,
            'score': sh.cell(i, 6).value,
            'cited': sh.cell(i, 3).value,
            'quartile': _get_q(q1, q2, q3, q4),
            'q1': q1,
            'q2': q2,
            'q3': q3,
            'q4': q4,
            'q0': [],
            'snip': sh.cell(i, 7).value,
            'sjr': sh.cell(i, 8).value,
            'top_10': 'YES' if sh.cell(i, 18).value == True else 'NO',
            'id': id,
            'url': sh.cell(i, 19).value,
            'open': sh.cell(i, 16).value.replace(' ', ''),
            'publisher': sh.cell(i, 14).value,
            'type': sh.cell(i, 15).value,
            'error': error,
        }

    print(f'... total records in file : {len(ids):-8d}')
    print(f'... skipped records       : {len(ids)-len(journals):-8d}')

    return journals


def export(journals, fpath='./data/journals.xlsx'):
    journals_out = Excel(fpath, description, _log, is_new=True)
    journals_out.create('journals', rewrite=True)
    journals_out.save('journals', journals)


def journals_parse():
    t = tpc()
    print(f'... Parsing')
    journals = build()
    num = len(journals.keys())
    print(f'>>> Parsed ({num:-8d} journals). Duration {tpc()-t:-.4f} sec')

    t = tpc()
    print(f'... Saving data')
    export(journals)
    print(f'>>> Data saved to excel file.   Duration {tpc()-t:-.4f} sec')


def _err(text, title=None):
    text = text + f' for "{title}"' if title else text
    print('WRN ' + text)
    return ' ' + text


def _get_q(q1, q2, q3, q4):
    if q1:
        return 'Q1'
    if q2:
        return 'Q2'
    if q3:
        return 'Q3'
    if q4:
        return 'Q4'


def _join_q(q1, q2):
    if q1 == 'Q1' or q2 == 'Q1':
        return 'Q1'
    if q1 == 'Q2' or q2 == 'Q2':
        return 'Q2'
    if q1 == 'Q3' or q2 == 'Q3':
        return 'Q3'
    if q1 == 'Q4' or q2 == 'Q4':
        return 'Q4'


def _log(text, kind='res'):
    print(kind.upper() + ' ' + text)


def _parse_codes(sh):
    codes = {}
    area = ''
    for i in range(2, CODES_MAX):
        numb = sh.cell(i, 1).value
        subj = sh.cell(i, 2).value
        if numb:
            codes[numb] = {
                'area': area,
                'subj': subj,
                'full': area + ' > ' + subj}
        elif subj:
            area = subj
    return codes


def _parse_issn(v):
    v = str(v) if v else None
    if v and len(v) < 8:
        v = '0' * (8-len(v)) + v
    if v and len(v) == 8:
        return v[:4] + '-' + v[4:]


if __name__ == '__main__':
    journals_parse()
