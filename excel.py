"""Скрипт для форматированной выгрузки данных в excel-файл.

Передаваемый словарь с описанием (desc) должен иметь вложенные словари, которые
имеют следующие поля:
- "name" - отображаемое в первой строке таблицы имя поля;
- "note" - заметка, которая будет добавлена к ячейке в первой строке;
- "size" - ширина столбца в таблице;
- "join" - разделитель (актуально, если соответствующие данные являются
списком, в этом случае список будет склеен в строку с данным разделителем);
- "dsbl" - если флаг задан, то столбец считается не активным и не вносится в
создаваемую таблицу;
- "align" - способ выравнивания данных для данного столбца (left, right или center). Если не задан, то используется глобальная переменная из "#opts";
- "color_func" - функция, которой передается значение элемента ячейки, и
возвращается цвет закраски ячейки;
- "color_head" - цвет закраски ячейки с заголовком;
- "separator" - если флаг задан, то справа от столбца будет нарисована рамка.

Дополнительно передаваемый словарь с описанием (desc) должен содержать служебное
поле "#opts", в котором содержится вложенный словарь с описанием дополнительных
параметров стилизации создаваемого excel-файла (см. образец).

Note:
    Если в строке данных есть поле с ключом "error", то для соотвествующей
    строки будет использована особая стилизация.

"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.utils import units
import os


NAMES_OWN = [
    'create', 'desc', 'file_path', 'load', 'load_all', 'log', 'save', 'wb']


class Excel():
    def __init__(self, file_path, desc, log, is_new=False):
        self.file_path = file_path
        self.desc = desc
        self.log = log

        if is_new:
            self.wb = openpyxl.Workbook()
            self.wb.remove(self.wb.active)
            # wb.active.title = sh_name
        else:
            self.wb = openpyxl.load_workbook(file_path)

        self.load_all()

    def create(self, name, rewrite=False):
        if name in self.wb.get_sheet_names():
            if not rewrite:
                self.log(f'Sheet "{name}" already exists', 'err')
                return
            self.wb.remove(self.wb[name])

        self.wb.create_sheet(title=name)

    def load(self, name, max_count=1000000):
        if not name in self.wb.get_sheet_names():
            self.log(f'Sheet "{name}" does not exist', 'err')
            return {}
        sh = self.wb[name]

        desc = self.desc.get(name)
        if desc is None:
            self.log(f'No description for sheet "{name}"', 'wrn')
            desc = {}

        def get_uid(name):
            for uid, item in (desc or {}).items():
                if item.get('name') == name:
                    return uid
            return name.lower().replace(' ', '_')

        def get_value(uid, v):
            item = (desc or {}).get(uid)
            if not item or not 'join' in item:
                return v
            return (v or '').split(item['join'])

        i = 1
        fields = []
        for j in range(1, max_count):
            name = sh.cell(i, j).value
            if name is None or not name or name == ' ':
                break
            fields.append(get_uid(name))

        table = {}
        for i in range(2, max_count):
            uid = sh.cell(i, 1).value
            if uid is None or not uid or uid == ' ':
                break

            row = {}
            for j, field in enumerate(fields, 1):
                value = sh.cell(i, j).value
                if value is not None and value != ' ':
                    row[field] = get_value(field, value)

            table[uid] = row

        return table

    def load_all(self):
        for name in self.desc.keys():
            if name in NAMES_OWN:
                self.log(f'Invalid sheet name "{name}"', 'err')
            if not name in self.wb.get_sheet_names():
                continue
            data = self.load(name)
            setattr(self, name, data)

    def save(self, name, data):
        if not name in self.wb.get_sheet_names():
            self.log(f'Sheet "{name}" does not exist', 'err')
            return
        sh = self.wb[name]

        desc = self.desc.get(name)
        if desc is None:
            self.log(f'No description for sheet "{name}"', 'err')
            return

        spec = desc.get('#opts', {})

        if spec.get('sheet_password'):
            sh.protection.sheet = True
            sh.protection.password = spec['sheet_password']

        if spec['freeze']:
            sh.freeze_panes = spec['freeze']

        if spec['fill']['tab']:
            sh.sheet_properties.tabColor = spec['fill']['tab']

        r = 1
        c = 1
        for uid, opts in desc.items():
            if uid == '#opts' or opts.get('dsbl'):
                continue

            cell = sh.cell(row=r, column=c)
            _prep(sh, cell, opts, spec, 'head')
            cell.value = opts.get('name', 'Unknown Name')
            c += 1

            if opts.get('note'):
                comment = cell.comment
                comment = Comment(opts['note'], spec.get('author', 'Author'))
                comment.width = spec['note']['width']
                comment.height = spec['note']['height']
                cell.comment = comment

        for r, item in enumerate(data.values(), 2):
            c = 1
            e = item.get('error')
            for uid, opts in desc.items():
                if uid == '#opts' or opts.get('dsbl'):
                    continue

                v = item.get(uid)
                if 'join' in opts:
                    v = opts['join'].join(v or [])
                is_empty = v is None or v == ''

                cell = sh.cell(row=r, column=c)
                _prep(sh, cell, opts, spec, 'data', v, e)
                if not is_empty:
                    cell.value = v
                c += 1

        self.wb.save(self.file_path)


def _prep(sh, cell, opts, spec, kind, v=None, e=None):
    is_empty = v is None or v == ''

    w = opts.get('size', 20)
    sh.column_dimensions[cell.column_letter].width = w

    h = spec['height'][kind]
    sh.row_dimensions[cell.row].height = h

    cell.font = Font(
        size=spec['font']['size'][kind],
        bold=spec['font']['bold'][kind],
        color=spec['font']['color']['error' if e else kind])

    align_hor = opts.get('align', spec['align']['hor'][kind])
    cell.alignment = Alignment(
        horizontal=align_hor,
        vertical=spec['align']['ver'][kind],
        wrap_text=spec['wrap'][kind])

    color = spec['fill'][kind]
    if is_empty and spec['fill']['empty']:
        color = spec['fill']['empty']
    if kind == 'head' and 'color_head' in opts:
        color = opts['color_head']
    if kind == 'data' and 'color_func' in opts:
        color_func = opts['color_func'](v)
        if color_func:
            color = color_func
    if color:
        cell.fill = PatternFill(
            fill_type='solid',
            start_color=color,
            end_color=color
        )

    border = Side(border_style='medium', color='FF000000')
    border_l = border if spec['border']['left'][kind] else None
    border_t = border if spec['border']['top'][kind] else None
    border_r = border if spec['border']['right'][kind] else None
    border_b = border if spec['border']['bottom'][kind] else None
    if opts.get('separator'):
        border_r = border
    cell.border = Border(
        left=border_l, top=border_t, right=border_r, bottom=border_b)
